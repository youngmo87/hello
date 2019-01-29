import csv, codecs
import openpyxl
from PIL import Image
from openpyxl.chart import (
    Reference,
    BarChart,
    ScatterChart,
    Series,
)

# Trythis 1번
book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "멜론top100리스트"

with open('./data/output.csv', 'r') as tempFile:  
    preData = csv.reader(tempFile)
    for i, list_info in enumerate(preData):
        sheet1.append(list_info)     
        for k in [1, 4, 5]:
            if i in [0, 101]:
                continue
            else:
                tmpCell = sheet1.cell(row=i+1, column=k)
                tmpCell.value = int(tmpCell.value)
                tmpCell.number_format              

book.save("./data/meltop100.xlsx")

# Trythis 2번
book = openpyxl.load_workbook("./data/meltop100.xlsx")
sheet2 = book.create_sheet()
sheet2.title = "멜론top100 앨범자켓"

for i in range(1, 101):
    # imgFile = './images/1.png'
    imgFile = "./images/" + str(i) + ".png"
    img=Image.open(imgFile)
    new_img = img.resize((140, 150))
    new_img.save("./images/" + str(i) + ".png")    
    img3 = openpyxl.drawing.image.Image(imgFile)
    col = ((i + 9) // 10)*2 - 1
    row = (7*i-6)%70
    sheet2.add_image(img3, '{}{}'.format(chr(col+64), row))
    
    # if i >= 1 and i <= 10:
    #     i = 'A' + str((i-1)*7+1)
    #     sheet2.add_image(img3, i)
    
    # elif i > 10 and i <=20:
    #     i = 'C' + str((i-11)*7+1)
    #     sheet2.add_image(img3, i)
    
    # elif i > 20  and i <=30:
    #     i = 'E' + str((i-21)*7+1)
    #     sheet2.add_image(img3, i)
 
    # elif i > 30 and i <=40:
    #     i = 'G' + str((i-31)*7+1)
    #     sheet2.add_image(img3, i)    
    
    # elif i > 40 and i <=50:
    #     i = 'I' + str((i-41)*7+1)
    #     sheet2.add_image(img3, i)    
    
    # elif i > 50 and i <=60:
    #     i = 'K' + str((i-51)*7+1)
    #     sheet2.add_image(img3, i)    
    
    # elif i > 60 and i <=70:
    #     i = 'M' + str((i-61)*7+1)
    #     sheet2.add_image(img3, i)    
    
    # elif i > 70 and i <=80:
    #     i = 'O' + str((i-71)*7+1)
    #     sheet2.add_image(img3, i)    
    
    # elif i > 80 and i <=90:
    #     i = 'Q' + str((i-81)*7+1)
    #     sheet2.add_image(img3, i)    
    
    # else:
    #     i = 'S' +str((i-91)*7+1)
    #     sheet2.add_image(img3, i)

book.save("./data/meltop100.xlsx")

# Trythis 3-1번
book = openpyxl.load_workbook("./data/meltop100.xlsx")
sheet3=book.create_sheet()
sheet3.title = "차트출력"

datax = Reference(sheet1, min_col=4, 
		min_row=2, max_col=4, max_row=11)
categs = Reference(sheet1, min_col=2,
				 min_row=2,max_row= 11)

chart1 = BarChart()
chart1.add_data(datax)
chart1.set_categories(categs)

chart1.legend = None  # 범례
chart1.varyColors = True
chart1.title = "상위 10위 좋아요수"
sheet3.add_chart(chart1, "A8")

# Trythis 3-2번
chart2 = ScatterChart()
chart2.style = 13
datax = Reference(sheet1, min_col=1, 
		min_row=2, max_row=11)

value=Reference(sheet1, min_col=5, min_row=1, max_row=11)
series = Series(value, datax, title_from_data=True)
chart2.series.append(series)
categs = Reference(sheet1, min_col=2,
				 min_row=2,max_row= 11)

# chart1.add_data(datax)
chart2.set_categories(categs)

chart2.legend = None  # 범례
chart2.varyColors = True
chart2.title = "상위 10위 좋아요차이수"
sheet3.add_chart(chart2, "K8")
book.save("./data/meltop100.xlsx")


